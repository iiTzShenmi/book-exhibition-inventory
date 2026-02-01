#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime guard
    load_workbook = None


HEADER_ALIASES = {
    "cabinet": {"平台", "櫃位", "櫃別", "陳列櫃", "cabinet", "cabinet_name"},
    "title": {"品名", "書名", "title", "book", "book_title"},
    "author": {"作者", "author", "author_name"},
}


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def find_header_row(ws, max_scan_rows=10):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers = [normalize(str(cell)) if cell is not None else "" for cell in row]
        if not any(headers):
            continue
        return row_idx, headers
    return None, []


def resolve_column(headers, names):
    for idx, header in enumerate(headers):
        if header in names:
            return idx
    return None


def main():
    parser = argparse.ArgumentParser(description="Import 2026 book list XLSX into inventory.csv")
    parser.add_argument("xlsx_path", help="Path to XLSX file")
    parser.add_argument("--sheet", help="Sheet name (default: active sheet)")
    parser.add_argument("--out", help="Output CSV path (default: Web/database/inventory.csv)")
    parser.add_argument("--backup", action="store_true", help="Backup existing inventory.csv before overwrite")
    args = parser.parse_args()

    if load_workbook is None:
        print("Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl", file=sys.stderr)
        return 2

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"XLSX not found: {xlsx_path}", file=sys.stderr)
        return 2

    base_dir = Path(__file__).resolve().parents[1]
    out_path = Path(args.out) if args.out else base_dir / "database" / "inventory.csv"

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    header_row, headers = find_header_row(ws)
    if not header_row:
        print("Failed to detect header row in the first 10 rows.", file=sys.stderr)
        return 2

    cabinet_idx = resolve_column(headers, HEADER_ALIASES["cabinet"])
    title_idx = resolve_column(headers, HEADER_ALIASES["title"])
    author_idx = resolve_column(headers, HEADER_ALIASES["author"])

    if cabinet_idx is None or title_idx is None:
        print("Missing required columns. Found headers:", headers, file=sys.stderr)
        print("Need columns: 平台 (cabinet), 品名 (title)", file=sys.stderr)
        return 2

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cab = normalize(str(row[cabinet_idx])) if row[cabinet_idx] is not None else ""
        title = normalize(str(row[title_idx])) if row[title_idx] is not None else ""
        if not cab or not title:
            continue
        author = ""
        if author_idx is not None and row[author_idx] is not None:
            author = normalize(str(row[author_idx]))
        rows.append([cab, title, "1", author] if author else [cab, title, "1"])

    if args.backup and out_path.exists():
        backup_dir = out_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"inventory_{ts}.csv"
        backup_path.write_bytes(out_path.read_bytes())
        print(f"Backup created: {backup_path}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
